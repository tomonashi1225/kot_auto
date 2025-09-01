import calendar
import configparser
import datetime
import tempfile

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.select import Select

from logger import get_module_logger

logger = get_module_logger(__name__, False)


EXCEL_FOLDER = "excels/"


def main():
    try:
        logger.info("config loading...")
        config = configparser.ConfigParser()
        config.read("./config.ini")
        logger.info("config loaded successfully.")

        # excelのロード
        print("年月を入力してください（yyyymm）")
        input_date = datetime.datetime.strptime(input(), "%Y%m")
        # input_date = datetime.datetime.strptime("202410", "%Y%m")
        today = datetime.datetime.now()
        month = input_date.strftime("%Y%m")
        logger.info("Excel loading...")
        wb = openpyxl.load_workbook(EXCEL_FOLDER + month + "_作業報告書.xlsx")
        sheet = wb["Sheet1"]
        logger.info("Excel loaded successfully.")

        # Webドライバーの初期化
        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument(f"--user-data-dir={tempfile.mkdtemp()}")
        options.binary_location = "/usr/bin/google-chrome"

        service = Service("/usr/local/bin/chromedriver")
        driver = webdriver.Chrome(service=service, options=options)
        driver.implicitly_wait(5)

        # ログイン
        logger.info("connecting browser...")
        driver.get("https://s2.kingtime.jp/admin")

        logger.info("loging in...")
        driver.find_element(By.ID, "login_id").send_keys(config["KOT"]["id"])
        driver.find_element(By.ID, "login_password").send_keys(config["KOT"]["pw"])
        driver.find_element(By.ID, "login_button").click()
        logger.info("login success.")

        # 出退勤入力開始
        logger.info("input start.")

        # 入力ページのチェック
        if input_date.year * 100 + input_date.month < today.year * 100 + today.month:
            diff = (today.year - input_date.year) * 12 + today.month - input_date.month
            for i in range(diff):
                driver.find_element(By.ID, "button_before_month").click()

        # 出退勤入力
        last_day = calendar.monthrange(input_date.year, input_date.month)[1]
        for i in range(last_day):
            start_time = sheet.cell(row=12 + i, column=5).value
            end_time = sheet.cell(row=12 + i, column=6).value

            if start_time is not None:

                menu_options = driver.find_elements(
                    By.CLASS_NAME, "htBlock-selectOther"
                )[i].find_elements(By.TAG_NAME, "option")
                edit_button = menu_options[1]
                edit_button.click()

                record_type_selectors = driver.find_elements(
                    By.CLASS_NAME, "htBlock-selectmenu"
                )
                Select(record_type_selectors[0]).select_by_visible_text("出勤")
                Select(record_type_selectors[1]).select_by_visible_text("退勤")

                record_time_editors = driver.find_elements(
                    By.CLASS_NAME, "recording_timestamp_time"
                )
                if record_time_editors[0].get_attribute("value") == "":
                    record_time_editors[0].send_keys(str(start_time)[:5])
                if record_time_editors[1].get_attribute("value") == "":
                    record_time_editors[1].send_keys(str(end_time)[:5])
                # 一回入力欄をクリックしないとなぜか打刻登録に反映されない
                record_time_editors[0].click()

                submit_button = driver.find_element(By.ID, "button_01")
                submit_button.click()

        # 勤怠確認済みとする
        check_button = driver.find_element(By.ID, "button_05")
        check_button.click()
        confirm_button = driver.find_element(By.ID, "button_01")
        confirm_button.click()

        logger.info("input complete.")

    except Exception as e:
        logger.error(e)
        raise


if __name__ == "__main__":
    main()
